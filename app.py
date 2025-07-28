import streamlit as st
import pandas as pd
import google.generativeai as genai
import json
from datetime import datetime
import re
import io
from fpdf import FPDF

# --- PAGE CONFIGURATION ---
st.set_page_config(
    page_title="MyMCMB AI Command Center",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- STYLING ---
st.markdown("""
<style>
    .stApp {
        background-color: #020617; /* Slate 950 */
        color: #e2e8f0; /* Slate 200 */
    }
    .main .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
    }
    h1, h2, h3 {
        color: #ffffff; 
        font-weight: 700;
    }
    .stButton>button {
        background-color: #2563eb; /* Blue 600 */
        color: white;
 
        border-radius: 0.5rem;
        border: none;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        transition: all 0.2s ease-in-out;
    }
    .stButton>button:hover {
        background-color: #1d4ed8; /* Blue 700 */
        transform: translateY(-2px);
    }
    .st-expander, .st-emotion-cache-18ni7ap {
        background-color: #0f172a; /* Slate 900 */
        border: 1px solid #1e293b; /* Slate 800 */
        border-radius: 0.5rem;
    }
    .stTextInput>div>div>input, .stNumberInput>div>div>input {
        background-color: #1e293b; /* Slate 800 */
        color: #e2e8f0;
    }
</style>
""", unsafe_allow_html=True)

# --- API & MODEL SETUP ---
try:
    GEMINI_API_KEY = st.secrets.get("GEMINI_API_KEY")
    APP_PASSWORD = st.secrets.get("APP_PASSWORD")
    if not GEMINI_API_KEY or not APP_PASSWORD:
        st.error("CRITICAL ERROR: API Keys or App Password are not configured. Please ensure they are set in your Streamlit Secrets in the correct TOML format.")
        st.stop()
        
    genai.configure(api_key=GEMINI_API_KEY)
    model = genai.GenerativeModel('gemini-1.5-flash')
except Exception as e:
    st.error(f"Could not configure AI models. Error: {e}")
    st.stop()

# --- SIDEBAR & NAVIGATION ---
st.sidebar.title("MyMCMB AI Command Center")
app_mode = st.sidebar.selectbox(
    "Select an AI Agent",
    ["Refinance Intelligence Center", "Admin Rate Panel", "Guideline & Product Chatbot", "Social Media Automation"]
)

# --- ROBUST SHARED FUNCTIONS ---
def clean_json_response(response_text):
    """Clean and parse AI-generated JSON responses with robust error handling."""
    try:
        # First, try to parse as-is
        return json.loads(response_text)
    except json.JSONDecodeError:
        try:
            # Remove potential markdown formatting
            cleaned = response_text.strip()
            if cleaned.startswith('```json'):
                cleaned = cleaned[7:]
            if cleaned.endswith('```'):
                cleaned = cleaned[:-3]
            cleaned = cleaned.strip()
            
            # Try parsing again
            return json.loads(cleaned)
        except json.JSONDecodeError:
            try:
                # Find JSON-like content between curly braces
                import re
                json_match = re.search(r'\{.*\}', cleaned, re.DOTALL)
                if json_match:
                    return json.loads(json_match.group())
                else:
                    # Return fallback structure
                    return {"outreach_options": []}
            except (json.JSONDecodeError, AttributeError):
                # Final fallback
                return {"outreach_options": []}

def clean_currency(value):
    if pd.isna(value): return 0.0
    try:
        return float(re.sub(r'[$,]', '', str(value)).strip())
    except (ValueError, TypeError):
        return 0.0

def clean_percentage(value):
    if pd.isna(value): return 0.0
    try:
        val = float(str(value).replace('%', '').strip())
        return val / 100 if val > 1 else val
    except (ValueError, TypeError):
        return 0.0

def calculate_new_pi(principal, annual_rate, term_years):
    try:
        principal = clean_currency(principal)
        annual_rate = clean_percentage(annual_rate)
        term_years = int(term_years)
        monthly_rate = annual_rate / 12
        num_payments = term_years * 12
     
        
        if monthly_rate <= 0 or num_payments <= 0: return 0.0
        payment = principal * (monthly_rate * (1 + monthly_rate)**num_payments) / ((1 + monthly_rate)**num_payments - 1)
        return round(payment, 2)
    except (ValueError, TypeError, ZeroDivisionError):
        return 0.0

def calculate_amortized_balance(principal, annual_rate, term_years, first_payment_date):
    try:
        principal = clean_currency(principal)
        annual_rate = clean_percentage(annual_rate)
        term_years = int(clean_currency(term_years))
        if pd.isna(first_payment_date): return principal
        
        first_payment = pd.to_datetime(first_payment_date)
        months_elapsed = (datetime.now().year - first_payment.year) * 12 + (datetime.now().month - first_payment.month)
        payments_made = max(0, months_elapsed)
        if payments_made == 0: return principal
        
        monthly_rate = annual_rate / 12
        total_payments = term_years * 12
        if monthly_rate <= 0: return principal * (1 - (payments_made / total_payments))
        
        balance = principal * ( ((1 + monthly_rate)**total_payments - (1 + monthly_rate)**payments_made) / ((1 + monthly_rate)**total_payments - 1) )
        return max(0, round(balance, 2))
    except Exception:
        return principal

# Mapping of common header variations to canonical names
COLUMN_ALIASES = {
    "Borrower First Name": ["borrower first name", "first name", "fname", "first"],
    "Borrower Last Name": ["borrower last name", "last name", "lname", "last"],
    "Current P&I Mtg Pymt": ["current p&i mtg pymt", "current payment", "pi payment", "current p&i"],
    "Original Property Value": ["original property value", "purchase price", "original home value", "home value"],
    "Total Original Loan Amount": ["total original loan amount", "original loan amount", "loan amount", "original loan balance"],
    "Current Interest Rate": ["current interest rate", "interest rate", "rate"],
    "Loan Term (years)": ["loan term (years)", "loan term", "term"],
    "First Pymt Date": ["first pymt date", "first payment date", "first payment"],
    "City": ["city", "property city", "borrower city"],
}

REQUIRED_COLUMNS = list(COLUMN_ALIASES.keys())

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename columns using COLUMN_ALIASES and return the DataFrame."""
    rename_map = {}
    lower_map = {c.lower(): c for c in df.columns}
    for target, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias.lower() in lower_map:
                rename_map[lower_map[alias.lower()]] = target
                break
    return df.rename(columns=rename_map)

# --- PDF EXPORT FUNCTION ---
class PDF(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, 'MyMCMB Refinance Opportunity Report', 0, 1, 'C')
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

    def chapter_title(self, title):
        self.set_font('Arial', 'B', 12)
        self.set_fill_color(240, 242, 246)
        self.cell(0, 10, title, 0, 1, 'L')
        self.ln(4)

    def chapter_body(self, body):
        self.set_font('Arial', '', 10)
        self.multi_cell(0, 5, body.encode('latin-1', 'replace').decode('latin-1'))
        self.ln()

def create_text_export(df_results):
    """Create a comprehensive text file export with all borrower data and outreach templates"""
    text_content = []
    text_content.append("=== MyMCMB AI Refinance Intelligence Report ===")
    text_content.append(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    text_content.append(f"Total Borrowers: {len(df_results)}")
    text_content.append("\n" + "="*80 + "\n")
    
    for index, row in df_results.iterrows():
        text_content.append(f"BORROWER #{index + 1}: {row['Borrower First Name']} {row.get('Borrower Last Name', '')}")
        text_content.append("-" * 50)
        text_content.append(f"City: {row.get('City', 'N/A')}")
        text_content.append(f"Current Monthly P&I: ${clean_currency(row['Current P&I Mtg Pymt']):,.2f}")
        text_content.append(f"Remaining Balance: ${row.get('Remaining Balance', 0):,.2f}")
        text_content.append(f"Estimated Home Value: ${row['Estimated Home Value']:,.2f}")
        text_content.append(f"Estimated LTV: {row['Estimated LTV']:.2f}%")
        text_content.append(f"Max Cash-Out Amount: ${row['Max Cash-Out Amount']:,.2f}")
        
        # Add refinance scenario details
        text_content.append("\nREFINANCE SCENARIOS:")
        scenario_cols = [c for c in df_results.columns if 'New P&I' in c or 'Savings' in c]
        for col in scenario_cols:
            if 'New P&I' in col:
                term = col.split('(')[1].split(')')[0]
                new_payment = row.get(col, 0)
                savings = row.get(f'Savings ({term})', 0)
                text_content.append(f"  {term}: New Payment: ${new_payment:.2f}, Monthly Savings: ${savings:.2f}")
        
        # Add AI-generated outreach options
        text_content.append("\nAI-GENERATED OUTREACH OPTIONS:")
        if row['AI_Outreach'] and row['AI_Outreach'].get('outreach_options'):
            for i, option in enumerate(row['AI_Outreach']['outreach_options']):
                text_content.append(f"\n  Option {i+1}: {option.get('title', 'N/A')}")
                text_content.append(f"  SMS: {option.get('sms', 'N/A')}")
                text_content.append(f"  Email: {option.get('email', 'N/A')}")
        else:
            text_content.append("  No outreach options generated.")
        
        text_content.append("\n" + "="*80 + "\n")
    
    return "\n".join(text_content)

def create_enhanced_pdf_report(df_results):
    """Create a comprehensive PDF report for all borrowers"""
    pdf = PDF()
    pdf.add_page()
    
    # Title page
    pdf.chapter_title("MyMCMB AI Refinance Intelligence Report")
    pdf.chapter_body(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    pdf.chapter_body(f"Total Borrowers: {len(df_results)}")
    
    for index, row in df_results.iterrows():
        pdf.add_page()
        pdf.chapter_title(f"Borrower: {row['Borrower First Name']} {row.get('Borrower Last Name', '')}")
        
        # Financial snapshot
        snapshot_text = f"""
FINANCIAL SNAPSHOT:
• City: {row.get('City', 'N/A')}
• Current Monthly P&I: ${clean_currency(row['Current P&I Mtg Pymt']):,.2f}
• Remaining Balance: ${row.get('Remaining Balance', 0):,.2f}
• Estimated Home Value: ${row['Estimated Home Value']:,.2f}
• Estimated LTV: {row['Estimated LTV']:.2f}%
• Max Cash-Out Amount: ${row['Max Cash-Out Amount']:,.2f}

REFINANCE SCENARIOS:
"""
        # Add refinance scenarios
        scenario_cols = [c for c in df_results.columns if 'New P&I' in c]
        for col in scenario_cols:
            if 'New P&I' in col:
                term = col.split('(')[1].split(')')[0]
                new_payment = row.get(col, 0)
                savings = row.get(f'Savings ({term})', 0)
                snapshot_text += f"• {term}: New Payment: ${new_payment:.2f}, Monthly Savings: ${savings:.2f}\n"
        
        pdf.chapter_body(snapshot_text)
        
        # Add outreach options
        if row['AI_Outreach'] and row['AI_Outreach'].get('outreach_options'):
            pdf.chapter_title("AI-Generated Outreach Options")
            for i, option in enumerate(row['AI_Outreach']['outreach_options']):
                pdf.chapter_body(f"Option {i+1}: {option.get('title', 'N/A')}")
                pdf.chapter_body(f"SMS: {option.get('sms', 'N/A')}")
                pdf.chapter_body(f"Email: {option.get('email', 'N/A')}")
                pdf.chapter_body("")
    
    output = pdf.output(dest='S')
    return bytes(output)

def create_pdf_report(data):
    pdf = PDF()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, f"For: {data['Borrower First Name']} {data.get('Borrower Last Name', '')}", 0, 1)
    
    pdf.chapter_title("Financial Snapshot")
    for key, value in data['snapshot'].items():
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(80, 8, f"{key}:", 1)
        pdf.set_font('Arial', '', 10)
        pdf.cell(0, 8, str(value), 1)
        pdf.ln()
    
    pdf.ln(10)
    pdf.chapter_title("AI-Generated Outreach Options")
    if data.get('outreach'):
        for option in data['outreach']:
            pdf.set_font('Arial', 'B', 11)
            pdf.cell(0, 10, option.get('title', 'N/A'), 0, 1)
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(0, 8, "SMS Template:", 0, 1)
            pdf.chapter_body(option.get('sms', 'N/A'))
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(0, 8, "Email Template:", 0, 1)
            pdf.chapter_body(option.get('email', 'N/A'))
            pdf.ln(5)

    output = pdf.output(dest='S')  # returns a bytearray
    return bytes(output)  # Streamlit expects raw PDF bytes for download

# --- ADMIN RATE PANEL ---
if app_mode == "Admin Rate Panel":
    st.title("Admin Rate Panel")
    password = st.text_input("Enter Admin Password", type="password")
    if password == APP_PASSWORD:
        st.success("Access Granted")
        st.write("Set the current mortgage rates. Leave a rate at 0 to disable that option for loan officers.")

        if 'rates' not in st.session_state:
            st.session_state.rates = {
                '30yr_fixed': 6.875, '25yr_fixed': 6.750, '20yr_fixed': 6.625, '15yr_fixed': 6.000, '10yr_fixed': 5.875,
                '7yr_arm': 7.125, '5yr_arm': 7.394, 'heloc': 8.500, 'no_cost_adj': 0.250
            }

        with st.form("rate_form"):
            st.subheader("Current Market Rates (%)")
            rates = st.session_state.rates
            col1, col2 = st.columns(2)
            with col1:
                rates['30yr_fixed'] = st.number_input("30-Year Fixed", value=rates['30yr_fixed'], format="%.3f")
                rates['25yr_fixed'] = st.number_input("25-Year Fixed", value=rates['25yr_fixed'], format="%.3f")
                rates['20yr_fixed'] = st.number_input("20-Year Fixed", value=rates['20yr_fixed'], format="%.3f")
                rates['15yr_fixed'] = st.number_input("15-Year Fixed", value=rates['15yr_fixed'], format="%.3f")
                rates['10yr_fixed'] = st.number_input("10-Year Fixed", value=rates['10yr_fixed'], format="%.3f")
            with col2:
                rates['7yr_arm'] = st.number_input("7/1 ARM", value=rates['7yr_arm'], format="%.3f")
                rates['5yr_arm'] = st.number_input("5/1 ARM", value=rates['5yr_arm'], format="%.3f")
                rates['heloc'] = st.number_input("HELOC Rate", value=rates['heloc'], format="%.3f")
                rates['no_cost_adj'] = st.number_input("No-Cost Adj.", value=rates['no_cost_adj'], format="%.3f", help="Amount to add for a no-cost option.")
            
            submitted = st.form_submit_button("Save Rates")
            if submitted:
                st.session_state.rates = rates
                st.success("Rates updated successfully!")
    elif password:
        st.error("Incorrect Password")

# --- REFINANCE INTELLIGENCE CENTER ---
elif app_mode == "Refinance Intelligence Center":
    st.title("Refinance Intelligence Center")
    st.markdown("### Choose how to input borrower data to generate hyper-personalized outreach plans.")
    
    # Data input method selection
    input_method = st.radio(
        "Select Data Input Method:",
        ["📁 Upload Excel File", "✏️ Manual Entry"],
        horizontal=True,
        help="Choose between uploading an Excel file or manually entering borrower information"
    )
    
    appreciation_rate = st.number_input(
        "Assumed Annual Home Appreciation Rate (%)",
        min_value=0.0,
        max_value=10.0,
        value=7.0,
        step=0.1,
        help="Used to estimate each borrower's current home value."
    )   
    
    if input_method == "📁 Upload Excel File":
        st.markdown("#### Upload Excel File")
        with st.expander("📊 Required Column Mapping"):
            mapping_df = pd.DataFrame({
                "Required Column": list(COLUMN_ALIASES.keys()),
                "Accepted Names": [", ".join(v) for v in COLUMN_ALIASES.values()],
            })
            st.dataframe(mapping_df, use_container_width=True)
        
        uploaded_file = st.file_uploader("Choose a borrower Excel file", type=['xlsx'])
    
    else:  # Manual Entry
        st.markdown("#### Manual Entry")
        st.markdown("Enter borrower information below. You can add multiple borrowers before generating the outreach plans.")
        
        # Quick Start Guide
        with st.expander("🚀 Quick Start Guide for Loan Officers"):
            st.markdown("""
            **How to Use Manual Entry Effectively:**
            
            1. **Gather Required Information**: You'll need the borrower's current payment, original loan details, and property info
            2. **Add Multiple Borrowers**: Use the form below to add each borrower one by one
            3. **Review Your List**: Check the borrowers you've added in the expandable sections
            4. **Generate Plans**: Click the "Generate AI Outreach Plans" button to create personalized messages
            5. **Download Reports**: Get Excel, PDF, or text formats of your outreach plans
            
            **Pro Tips:**
            - **Interest Rate**: Enter as percentage (e.g., 7.250 for 7.25%)
            - **Payment Amount**: Enter the current P&I payment only (not including taxes/insurance)
            - **Property Value**: Use the original purchase price from when they bought
            - **First Payment Date**: This helps calculate how much they still owe
            
            **Time Saver**: If you have an Excel file with multiple borrowers, use the "Upload Excel File" option instead!
            """)
        
        # Initialize session state for manual entries
        if 'manual_borrowers' not in st.session_state:
            st.session_state.manual_borrowers = []
        
        with st.form("manual_entry_form"):
            st.subheader("Add New Borrower")
            col1, col2 = st.columns(2)
            
            with col1:
                first_name = st.text_input("Borrower First Name*", placeholder="John")
                last_name = st.text_input("Borrower Last Name*", placeholder="Smith")
                city = st.text_input("City*", placeholder="Nashville")
                current_payment = st.number_input("Current P&I Monthly Payment*", min_value=0.0, step=10.0, format="%.2f", placeholder=1500.00)
                original_value = st.number_input("Original Property Value*", min_value=0.0, step=1000.0, format="%.2f", placeholder=300000.00)
            
            with col2:
                original_loan = st.number_input("Total Original Loan Amount*", min_value=0.0, step=1000.0, format="%.2f", placeholder=240000.00)
                current_rate = st.number_input("Current Interest Rate (%)*", min_value=0.0, max_value=20.0, step=0.125, format="%.3f", placeholder=7.250)
                loan_term = st.number_input("Loan Term (years)*", min_value=1, max_value=50, step=1, value=30)
                first_payment = st.date_input("First Payment Date*", help="Date of the first mortgage payment")
            
            st.markdown("*Required fields")
            
            submit_borrower = st.form_submit_button("➕ Add Borrower", use_container_width=True)
            
            if submit_borrower:
                # Validate required fields
                if not all([first_name, last_name, city, current_payment > 0, original_value > 0, 
                           original_loan > 0, current_rate > 0, loan_term > 0, first_payment]):
                    st.error("Please fill in all required fields.")
                else:
                    # Add borrower to session state
                    borrower_data = {
                        "Borrower First Name": first_name,
                        "Borrower Last Name": last_name,
                        "City": city,
                        "Current P&I Mtg Pymt": current_payment,
                        "Original Property Value": original_value,
                        "Total Original Loan Amount": original_loan,
                        "Current Interest Rate": current_rate / 100,  # Convert percentage to decimal
                        "Loan Term (years)": loan_term,
                        "First Pymt Date": first_payment
                    }
                    st.session_state.manual_borrowers.append(borrower_data)
                    st.success(f"Added {first_name} {last_name} to the processing queue!")
                    st.rerun()
        
        # Display current borrowers
        if st.session_state.manual_borrowers:
            st.subheader(f"📋 Borrowers Ready for Processing ({len(st.session_state.manual_borrowers)})")
            
            # Show a quick summary
            total_current_payments = sum(b['Current P&I Mtg Pymt'] for b in st.session_state.manual_borrowers)
            avg_payment = total_current_payments / len(st.session_state.manual_borrowers)
            
            col1, col2, col3 = st.columns(3)
            col1.metric("Total Borrowers", len(st.session_state.manual_borrowers))
            col2.metric("Avg Current Payment", f"${avg_payment:,.2f}")
            col3.metric("Total Monthly Payments", f"${total_current_payments:,.2f}")
            
            for i, borrower in enumerate(st.session_state.manual_borrowers):
                with st.expander(f"{borrower['Borrower First Name']} {borrower['Borrower Last Name']} - {borrower['City']}"):
                    col1, col2, col3 = st.columns([3, 1, 1])
                    with col1:
                        st.write(f"**Current P&I:** ${borrower['Current P&I Mtg Pymt']:,.2f}")
                        st.write(f"**Original Property Value:** ${borrower['Original Property Value']:,.2f}")
                        st.write(f"**Current Rate:** {borrower['Current Interest Rate']*100:.3f}%")
                        st.write(f"**First Payment:** {borrower['First Pymt Date']}")
                    with col2:
                        if st.button("✏️ Edit", key=f"edit_{i}"):
                            st.info("Edit functionality: Please remove and re-add the borrower for now.")
                    with col3:
                        if st.button("🗑️ Remove", key=f"remove_{i}"):
                            st.session_state.manual_borrowers.pop(i)
                            st.rerun()
            
            col1, col2 = st.columns([1, 1])
            with col1:
                if st.button("🗑️ Clear All Borrowers", type="secondary"):
                    st.session_state.manual_borrowers = []
                    st.rerun()
            with col2:
                if st.button("📊 Preview Summary", type="secondary"):
                    st.info(f"Ready to process {len(st.session_state.manual_borrowers)} borrowers with a combined monthly payment volume of ${total_current_payments:,.2f}")
        
        # Set uploaded_file to None for manual entry path
        uploaded_file = None
    
    # Determine data source and prepare for processing
    df_original = None
    data_source = None
    
    if uploaded_file:
        try:
            df_original = pd.read_excel(uploaded_file, engine='openpyxl')
            df_original = normalize_columns(df_original)
            missing = [c for c in REQUIRED_COLUMNS if c not in df_original.columns]
            if missing:
                st.error(
                    f"Missing required columns after header normalization: {', '.join(missing)}."
                )
                st.stop()
            data_source = f"Excel file '{uploaded_file.name}'"
            st.success(
                f"Successfully loaded {len(df_original)} borrowers from '{uploaded_file.name}'."
            )
        except Exception as e:
            st.error(f"An error occurred while processing the file. Please check your Excel sheet for correct column names and data types. Error: {e}")
    
    elif input_method == "✏️ Manual Entry" and st.session_state.get('manual_borrowers'):
        try:
            df_original = pd.DataFrame(st.session_state.manual_borrowers)
            data_source = "manual entry"
            st.success(
                f"Successfully prepared {len(df_original)} borrowers from manual entry."
            )
        except Exception as e:
            st.error(f"An error occurred while processing manual entries. Error: {e}")
    
    # Process data if available
    if df_original is not None and len(df_original) > 0:            
        if st.button("🚀 Generate AI Outreach Plans"):
            try:
                with st.spinner("Initiating AI Analysis... This will take a few moments."):
                    df = df_original.copy()
                    rates = st.session_state.get('rates', {'30yr_fixed': 6.875, '20yr_fixed': 6.625, '15yr_fixed': 6.000, '10yr_fixed': 5.875, '25yr_fixed': 6.750, '7yr_arm': 7.125, '5yr_arm': 7.394, 'heloc': 8.500, 'no_cost_adj': 0.250})

                    progress_bar = st.progress(0, text="Calculating financial scenarios...")
                df['Remaining Balance'] = df.apply(lambda row: calculate_amortized_balance(row.get('Total Original Loan Amount'), row.get('Current Interest Rate'), row.get('Loan Term (years)'), row.get('First Pymt Date')), axis=1)
                df['Months Since First Payment'] = df['First Pymt Date'].apply(lambda x: max(0, (datetime.now().year - pd.to_datetime(x).year) * 12 + (datetime.now().month - pd.to_datetime(x).month)) if pd.notna(x) else 0)
                df['Estimated Home Value'] = df.apply(
                    lambda row: round(
                        clean_currency(row.get('Original Property Value', 0))
                        * ((1 + appreciation_rate / 100) ** (row['Months Since First Payment'] / 12)),
                        2,
                    ),
                    axis=1,
                )
                df['Estimated LTV'] = (
                    (df['Remaining Balance'] / df['Estimated Home Value'])
                ).fillna(0).replace([float('inf'), -float('inf')], 0).round(4)
                df['Max Cash-Out Amount'] = (df['Estimated Home Value'] * 0.80) - df['Remaining Balance']
                df['Max Cash-Out Amount'] = df['Max Cash-Out Amount'].apply(lambda x: max(0, round(x, 2)))

                
                rate_terms = [
                    ('30yr', '30yr_fixed', 30),
                    ('25yr', '25yr_fixed', 25),
                    ('20yr', '20yr_fixed', 20),
                    ('15yr', '15yr_fixed', 15),
                    ('10yr', '10yr_fixed', 10),
                    ('7yrARM', '7yr_arm', 30),
                    ('5yrARM', '5yr_arm', 30),
                ]

                for term, rate_key, years in rate_terms:
                    rate = rates.get(rate_key, 0) / 100
                    if rate > 0:
                        df[f'New P&I ({term})'] = df.apply(lambda row: calculate_new_pi(row['Remaining Balance'], rate, years), axis=1)
                        df[f'Savings ({term})'] = df.apply(lambda row: clean_currency(row['Current P&I Mtg Pymt']) - row[f'New P&I ({term})'], axis=1)
                # Heloc interest-only payment estimate
               
                heloc_rate = rates.get('heloc', 0) / 100
                if heloc_rate > 0:
                    
                    df['HELOC Payment (interest-only)'] = df['Max Cash-Out Amount'] * (heloc_rate / 12)
                
                # Initialize outreach results with proper length to prevent mismatch
                outreach_results = [{"outreach_options": []} for _ in range(len(df))]
                
               
                for i, row in df.iterrows():
                    try:
                        progress_bar.progress((i + 1) / len(df), text=f"Generating AI outreach for {row['Borrower First Name']} {row.get('Borrower Last Name', '')}...")

                        current_payment = clean_currency(row['Current P&I Mtg Pymt'])
                        new_rate = rates.get('30yr_fixed', 0) / 100
                        no_cost_adj = rates.get('no_cost_adj', 0)
                        try:
                            max_loan_for_same_payment = (current_payment * (((1 + new_rate/12)**360) - 1)) / ((new_rate/12) * (1 + new_rate/12)**360) if new_rate > 0 else 0
                            cash_out_same_payment = max(0, max_loan_for_same_payment - row['Remaining Balance'])
                        except (ZeroDivisionError, ValueError):
                            cash_out_same_payment = 0
                        
                        prompt = f"""
                        You are an expert mortgage loan officer assistant for MyMCMB. You previously helped {row['Borrower First Name']} close their loan and are now following up as their trusted loan officer. The tone must be professional yet warm and personal, acknowledging your past relationship.

                        **Borrower's Financial Snapshot:**
                        - Property City: {row.get('City', 'their city')}
                        - Current Monthly P&I: ${clean_currency(row['Current P&I Mtg Pymt']):.2f}
                        - Estimated Home Value: ${row['Estimated Home Value']:.2f}
                        - Estimated LTV: {row['Estimated LTV']:.2f}%

                        **Calculated Refinance Scenarios:**
                        1.  **30-Year Fixed:** New Payment: ${row.get('New P&I (30yr)', 0):.2f}, Monthly Savings: ${row.get('Savings (30yr)', 0):.2f}
                        2.  **15-Year Fixed:** New Payment: ${row.get('New P&I (15yr)', 0):.2f}, Monthly Savings: ${row.get('Savings (15yr)', 0):.2f}
                        3.  **Max Cash-Out:** You can offer up to ${row['Max Cash-Out Amount']:.2f} in cash.
                        4.  **"Same Payment" Cash-Out:** You can offer approx. ${cash_out_same_payment:.2f} in cash while keeping their payment nearly the same.

                        **Task:**
                        Return a JSON object with a key named 'outreach_options'. That list should contain four distinct outreach options. Each option must have a 'title', a concise 'sms' template, and a professional 'email' template. 
                        
                        **Critical Guidelines for Maximum Conversion:**
                        - Use relationship language: "I hope you and your family are doing well", "It's been a while since we closed your loan", "As your previous loan officer", "I wanted to personally reach out", "Following up with my preferred clients"
                        - Mention they qualify for a **no-cost refinance** option where all fees are covered by lender for roughly +{no_cost_adj:.3f}% to the rate
                        - Create urgency without being pushy: "rates may change", "limited time opportunity", "current market conditions"
                        - Include specific dollar amounts and savings to make it tangible
                        - Make messages sound authentic and conversational, not sales-y
                        - Include genuine concern for their financial wellbeing
                        - End with a clear, specific call-to-action
                        - Keep SMS under 160 characters for best delivery
                        
                        **Four Required Outreach Approaches:**
                        1.  **"Significant Savings Alert"**: Focus on the 30-year option's direct monthly savings. Emphasize the immediate financial relief. Use language like "I wanted to personally update you on some exciting refinance opportunities that could put money back in your pocket every month"
                        
                        2.  **"Aggressive Payoff Plan"**: Focus on the 15-year option, highlighting owning their home faster and building wealth. Use language like "I've been thinking about your financial goals and found a way to help you own your home outright years earlier"
                        
                        3.  **"Leverage Your Equity"**: Focus on the maximum cash-out option for home improvements, debt consolidation, or investments. Use language like "I noticed your home value has grown significantly since we closed your loan - this creates some incredible opportunities for you"
                        
                        4.  **"Cash with No Payment Shock"**: Focus on the 'same payment' cash-out option. Use language like "I found a way to get you cash for your needs without changing your monthly payment - this is probably the most popular option with my clients"
                        
                        **Personalization Requirements:**
                        - For one of the emails, mention a positive local trend in {row.get('City', 'their area')} (like real estate growth, local economy, new developments)
                        - Include their first name in all messages
                        - Reference the specific dollar amounts from their scenario
                        - Use "we" language to build partnership ("let's review", "we can explore")
                        """
                        
                 
                        try:
                            response = model.generate_content(
                    
                 
                                prompt,
                                generation_config=genai.types.GenerationConfig(
                                    
                                    response_mime_type="application/json"
                                ),
                            )
                            
                            # Use robust JSON parsing

                            data = clean_json_response(response.text)
                            
                            # Validate the response structure
                            if 'outreach_options' not in data or not isinstance(data['outreach_options'], list):
                                raise ValueError("AI response missing required 'outreach_options' key or invalid structure")
                            
                            # Ensure we have at least some outreach options
                            if len(data['outreach_options']) == 0:
                                raise ValueError("AI response contains empty outreach_options")
                            
                            # Validate each outreach option has required fields
                            for option in data['outreach_options']:
                                if not isinstance(option, dict) or 'title' not in option or 'sms' not in option or 'email' not in option:
                                    raise ValueError("Invalid outreach option structure")
                            
                            outreach_results[i] = data
                            
                        except Exception as api_error:
                            st.warning(
                                f"AI content generation failed for {row['Borrower First Name']} {row.get('Borrower Last Name', '')}. Error: {api_error}"
                            )
                            # Keep the default empty structure that was pre-initialized
                            
                    except Exception as row_error:
                        st.error(
                            f"Error processing borrower {row.get('Borrower First Name', 'Unknown')}: {row_error}")
                        # Keep the default empty structure that was pre-initialized

                
                # ASSIGN OUTREACH RESULTS AFTER THE LOOP - Length is guaranteed to match
                df['AI_Outreach'] = outreach_results
                st.session_state.df_results = df
                st.success("Analysis complete! View the outreach plans below.")
            except Exception as e:
                st.error(f"An
                error occurred while processing the data. Error: {e}")
    
    elif input_method == "✏️ Manual Entry" and not st.session_state.get('manual_borrowers'):
        st.info("👆 Please add borrowers using the form above to get started.")
    elif input_method == "📁 Upload Excel File" and not uploaded_file:
        st.info("👆 Please upload an Excel file to get started.")
    
    if 'df_results' in st.session_state:
        st.markdown("---")
        st.header("Generated Outreach Blueprints")
        df_results = st.session_state.df_results
        
        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            export_df = df_results.copy()
            for i, row in export_df.iterrows():
                if row['AI_Outreach'] and row['AI_Outreach'].get('outreach_options'):
                    for j, option in enumerate(row['AI_Outreach']['outreach_options']):
                        export_df.at[i, f'Option_{j+1}_Title'] = option.get('title')
                  
                        export_df.at[i, f'Option_{j+1}_SMS'] = option.get('sms')
                        export_df.at[i, f'Option_{j+1}_Email'] = option.get('email')
            export_df = export_df.drop('AI_Outreach', axis=1)
            
            preferred_order = [
                'Borrower First Name', 'Borrower Last Name', 'City',
                
                'Current P&I Mtg Pymt', 'Remaining Balance',
                'Estimated Home Value', 'Estimated LTV', 'Max Cash-Out Amount',
                'HELOC Payment (interest-only)',

            ]
            scenario_cols = sorted([c for c in export_df.columns if 'New P&I' in c or 'Savings' in c])
            cols_in_order = [c for c in preferred_order if c in export_df.columns] + scenario_cols
            cols_in_order += [c for c in export_df.columns if c not in cols_in_order]
            expor
            t_df = export_df[cols_in_order]
            
            export_df.to_excel(writer, index=False, sheet_name='AI_Outreach_Plan')
            worksheet = writer.book['AI_Outreach_Plan']
            worksheet.freeze_panes = 'A2'
            
            # Enhanced styling with color coding
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            # Header styling
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            # Data styling
            savings_fill = PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid')  # Light green for savings
            cashout_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Light yellow for cash-out
            contact_fill = PatternFill(start_color='E1D5E7', end_color='E1D5E7', fill_type='solid')  # Light purple for contact info
            
            # Apply header styling
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply data styling based on column type
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    col_name = worksheet.cell(row=1, column=cell.column).value
                    
             
                    if col_name and 'Savings' in col_name:
                        cell.fill = savings_fill
                        if isinstance(cell.value, (int, float)) and cell.value > 0:
                            cell.font = Font(color='006100', bold=True)  # Dark green for positive savings
                    elif col_name and ('Cash-Out' in col_name or 'HELOC' in col_name):
                        cell.fill = cashout_fill
                        if isinstance(cell.value, (int, float)) and cell.value > 0:
                            cell.font = Font(color='9C6500', bold=True)  # Dark yellow for cash amounts
                    elif col_name and ('Name' in col_name or 'City' in col_name):
                        cell.fill = contact_fill
                        cell.font = Font(color='5B2C6F', bold=True)  # Purple for contact info
            
            # Auto-adjust column widths
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(50, length + 2)
            
            # Summary sheet with enhanced formatting
            savings_cols = [c for c in scenario_cols if c.startswith('Savings')]
            summary_df = export_df[
                ['Borrower First Name', 'Borrower Last Name', 'Estimated Home Value', 'Estimated LTV', 'Max Cash-Out Amount'] + savings_cols

            ].copy()

            summary_df['Best Savings'] = summary_df[savings_cols].max(axis=1)
            summary_df['Best Option'] = summary_df[savings_cols].idxmax(axis=1).str.extract(r'\((.*)\)')
            summary_df.to_excel(writer, index=False, sheet_name='Summary')
            
            summary_ws = writer.book['Summary']
            summary_ws.freeze_panes = 'A2'
            
            # Apply enhanced styling to summary sheet
            for cell in summary_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.
                alignment = Alignment(horizontal='center', vertical='center')

            
            # Highlight best savings
            for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row):
                best_savings_cell = row[-2]  # Best Savings column
                if isinstance(best_savings_cell.value, (int, float)) and best_savings_cell.value > 100:
                    best_savings_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')  # Bright green
                    best_savings_cell.font = Font(color='FFFFFF', bold=True)
            
            for column_cells in summary_ws.columns:
                length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                summary_ws.column_dimensions[column_cells[0].column_letter].width = min(50, length + 2)
        
        # Enhanced Export Options
        st.subheader("📥 Export Options")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.download_button(
                label="📊 Download Enhanced Excel Report",
                data=output_buffer.getvalue(),
                file_name="AI_Outreach_Plan.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Complete Excel report with color coding and multiple sheets"
            )
        
        with col2:
            text_export = create_text_export(df_results)
            st.download_button(
                label="📄 Download Text Report",
                data=text_export,
                file_name="AI_Outreach_Plan.txt",
                mime="text/plain",
                help="Comprehensive text file with all borrower data and outreach templates"
            )

        
        with col3:
            pdf_export = create_enhanced_pdf_report(df_results)
            st.download_button(
                label="📋 Download Complete PDF Report",
                data=pdf_export,
                file_name="AI_Outreach_Complete_Report.pdf",
                mime="application/pdf",
                help="Professional PDF report with all borrowers and outreach plans"
            )
        
        for index, row in df_results.iterrows():
            with st.expander(f"👤 **{row['Borrower First Name']} {row.get('Borrower Last Name', '')}** | Max Savings: **${row.get('Savings (30yr)', 0):.2f}/mo**"):
                st.subheader("Financial Snapshot")
                col1, col2, col3 = st.columns(3)
                col1.metric("Current P&I", f"${clean_currency(row['Current P&I Mtg Pymt']):,.2f}")
                col2.metric("Est. New P&I (30yr)", f"${row.get('New P&I (30yr)', 0):.2f}", delta=f"{-row.get('Savings (30yr)', 0):.2f}")
                col3.metric("Max Cash-Out", f"${row['Max Cash-Out Amount']:,.2f}")
                
                st.subheader("AI-Generated Outreach Options")
                if row['AI_Outreach'] and row['AI_Outreach'].get('outreach_options'):
                    pdf_data = {
                        "Borrower First Name": row['Borrower First Name'],
                        "Borrower Last Name": row.get('Borrower Last Name', ''),
                        "snapshot": {
                            "Current P&I": f"${clean_currency(row['Current P&I Mtg Pymt']):,.2f}",
                            "Est. New P&I (30yr)": f"${row.get('New P&I (30yr)', 0):.2f}",
                            "Max Cash-Out": f"${row['Max Cash-Out Amount']:,.2f}"
                        },
                        "outreach": row['AI_Outreach']['outreach_options']
                    }
                    st.download_button(
                        label="📄 Download PDF Summary",
                        data=create_pdf_report(pdf_data),
                        file_name=f"{row['Borrower First Name']}_Report.pdf",
                        mime="application/pdf",
                        key=f"pdf_{index}"
                    )
                    for option in row['AI_Outreach']['outreach_options']:
                        st.markdown(f"#### {option.get('title', 'Outreach Option')}")
                        st.text_area("SMS", value=option.get('sms'), height=100, key=f"sms_{index}_{option.get('title')}", help="Copy this template for SMS outreach.")
                        st.text_area("Email", value=option.get('email'), height=200, key=f"email_{index}_{option.get('title')}", help="Copy this template for email outreach.")
                else:
                    st.warning("Could not generate outreach content for this borrower.")
    
    # --- GUIDELINE & PRODUCT CHATBOT ---
elif app_mode == "Guideline & Product Chatbot":
    st.title("Guideline & Product Chatbot")
    st.markdown("### Ask questions about mortgage guidelines, products, and underwriting requirements.")
    
    # Initialize chat history
    if 'chat_history' not in st.session_state:
        st.session_state.chat_history = []
    
    # Knowledge base (can be expanded with document upload in the future)
    MORTGAGE_KNOWLEDGE = """
    MORTGAGE UNDERWRITING GUIDELINES AND PRODUCTS:
    
    CONVENTIONAL LOANS:
    - Minimum credit score: 620 for most programs
    - Down payment: As low as 3% for first-time homebuyers
    - Maximum DTI: 45% typically, 50% with compensating factors
    - PMI required for loans over 80% LTV
    - Loan limits vary by county (2024: $766,550 in most areas, up to $1,149,825 in high-cost areas)
    
    FHA LOANS:
    - Minimum credit score: 580 for 3.5% down, 500 for 10% down
    - Down payment: As low as 3.5%
    - Maximum DTI: 57% typically
    - MIP required for most loans
    - Loan limits: Generally lower than conventional
    
    VA LOANS:
    - No down payment required
    - No PMI required
    - Minimum credit score varies by lender (typically 620+)
    - Certificate of Eligibility required
    - Funding fee applies (can be financed)
    
    USDA LOANS:
    - No down payment required
    - Income limits apply
    - Property must be in eligible rural/suburban area
    - Guarantee fee required
    
    JUMBO LOANS:
    - Loan amounts above conforming limits
    - Typically require higher credit scores (700+)
    - Larger down payments often required
    - More stringent debt-to-income requirements
    
    REFINANCE OPTIONS:
    - Rate & Term Refinance: Lower rate or change terms
    - Cash-Out Refinance: Access home equity (typically max 80% LTV)
    - Streamline Refinance: Simplified process for existing customers
    - HARP/FHFA Programs: For underwater mortgages
    
    COMMON UNDERWRITING FACTORS:
    - Employment history (2+ years stable)
    - Asset verification (2+ months statements)
    - Property appraisal and title work
    - Debt-to-income ratios
    - Credit history and score
    - Reserves (2+ months PITI for investment properties)
    
    SPECIAL PROGRAMS:
    - First-time homebuyer programs
    - Down payment assistance programs
    - Energy-efficient mortgage programs
    - Renovation loans (203k, CHOICEReno)
    - Physician loans (high DTI tolerance)
    - Bank statement loans for self-employed
    """
    
    def get_chatbot_response(question, chat_history):
        """Generate response using the mortgage knowledge base"""
        try:
            # Create context from recent chat history
            context = ""
            if chat_history:
                recent_history = chat_history[-6:]  # Last 3 exchanges
                for entry in recent_history:
                    context += f"Previous Q: {entry['question']}\nPrevious A: {entry['answer']}\n\n"
            
            prompt = f"""
            You are a knowledgeable mortgage loan officer assistant for MyMCMB. You help loan officers and clients understand mortgage guidelines, products, and underwriting requirements.
            
            KNOWLEDGE BASE:
            {MORTGAGE_KNOWLEDGE}
            
            RECENT CONVERSATION CONTEXT:
            {context}
            
            USER QUESTION: {question}
            
            Please provide a helpful, accurate response based on the knowledge base. If the question is outside your knowledge area, suggest they contact their loan officer directly. Be professional, clear, and specific with rates, percentages, and requirements when applicable.
            
            If someone asks about current rates, remind them that rates change daily and they should contact their loan officer for current pricing.
            """
            
            response = model.generate_content(prompt)
            return response.text
            
        except Exception as e:
            return f"I apologize, but I'm having trouble generating a response right now. Please contact your loan officer directly for assistance. Error: {str(e)}"
    
    # Chat interface
    st.markdown("#### 💬 Chat with the Mortgage Expert")
    
    # Display chat history
    chat_container = st.container()
    with chat_container:
        for i, entry in enumerate(st.session_state.chat_history):
            # User message
            st.markdown(f"""
            <div style="background-color: #2563eb; color: white; padding: 10px; border-radius: 10px; margin: 5px 0; max-width: 80%; margin-left: auto;">
                <strong>You:</strong> {entry['question']}
            </div>
            """, unsafe_allow_html=True)
            
            # Bot response
            st.markdown(f"""
            <div style="background-color: #1e293b; color: #e2e8f0; padding: 10px; border-radius: 10px; margin: 5px 0; max-width: 80%;">
                <strong>🏠 Mortgage Expert:</strong> {entry['answer']}
            </div>
            """, unsafe_allow_html=True)
    
    # Input form
    with st.form("chat_form", clear_on_submit=True):
        user_question = st.text_area(
            "Ask your mortgage question:",
            placeholder="e.g., What are the current FHA loan requirements? or Can I qualify for a conventional loan with a 650 credit score?",
            height=100
        )
        
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            submit_chat = st.form_submit_button("💬 Ask Question", use_container_width=True)
        
        if submit_chat and user_question.strip():
            with st.spinner("Getting answer from mortgage expert..."):
                response = get_chatbot_response(user_question, st.session_state.chat_history)
                
                # Add to chat history
                st.session_state.chat_history.append({
                    'question': user_question,
                    'answer': response,
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                
                st.rerun()
    
    # Quick question buttons
    st.markdown("#### 🔥 Popular Questions")
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("💳 Credit Score Requirements", use_container_width=True):
            question = "What are the minimum credit score requirements for different loan types?"
            response = get_chatbot_response(question, st.session_state.chat_history)
            st.session_state.chat_history.append({
                'question': question,
                'answer': response,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            st.rerun()
        
        if st.button("🏠 Down Payment Options", use_container_width=True):
            question = "What are the minimum down payment requirements for first-time homebuyers?"
            response = get_chatbot_response(question, st.session_state.chat_history)
            st.session_state.chat_history.append({
                'question': question,
                'answer': response,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            st.rerun()
    
    with col2:
        if st.button("📊 Debt-to-Income Limits", use_container_width=True):
            question = "What are the maximum debt-to-income ratios for different loan programs?"
            response = get_chatbot_response(question, st.session_state.chat_history)
            st.session_state.chat_history.append({
                'question': question,
                'answer': response,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            st.rerun()
        
        if st.button("🔄 Refinance Options", use_container_width=True):
            question = "What refinance options are available and what are the requirements?"
            response = get_chatbot_response(question, st.session_state.chat_history)
            st.session_state.chat_history.append({
                'question': question,
                'answer': response,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            st.rerun()
    
    # Clear chat option
    if st.session_state.chat_history:
        st.markdown("---")
        if st.button("🗑️ Clear Chat History"):
            st.session_state.chat_history = []
            st.rerun()
        
        # Export chat option
        if len(st.session_state.chat_history) > 0:
            chat_export = "\n".join([
                f"Q: {entry['question']}\nA: {entry['answer']}\nTime: {entry['timestamp']}\n{'-'*50}"
                for entry in st.session_state.chat_history
            ])
            
            st.download_button(
                label="📄 Export Chat History",
                data=chat_export,
                file_name=f"mortgage_chat_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                mime="text/plain"
            )

elif app_mode == "Social Media Automation":
    st.title("Social Media Automation")
    st.markdown("### Generate engaging social media content for mortgage professionals")
    
    # Content type selection
    content_type = st.selectbox(
        "Select Content Type:",
        [
            "📊 Market Update Post",
            "💡 Educational Content", 
            "🏠 Home Buying Tips",
            "📈 Rate Alert",
            "🎯 First-Time Buyer Guide",
            "💰 Refinance Opportunity",
            "📝 Client Testimonial Template",
            "🎉 Closing Celebration"
        ]
    )
    
    # Platform selection
    platforms = st.multiselect(
        "Select Social Media Platforms:",
        ["📘 Facebook", "📸 Instagram", "🐦 Twitter/X", "💼 LinkedIn", "🎵 TikTok"],
        default=["📘 Facebook", "📸 Instagram"]
    )
    
    if not platforms:
        st.warning("Please select at least one social media platform.")
        st.stop()
    
    # Additional inputs based on content type
    col1, col2 = st.columns(2)
    
    with col1:
        if "Market Update" in content_type or "Rate Alert" in content_type:
            current_rate = st.number_input("Current 30-Year Rate (%)", min_value=0.0, max_value=15.0, value=7.0, step=0.125, format="%.3f")
            rate_change = st.selectbox("Rate Movement", ["📈 Increased", "📉 Decreased", "➡️ Remained Stable"])
        
        if "Educational" in content_type or "Tips" in content_type:
            focus_area = st.selectbox(
                "Focus Area:",
                ["Credit Score Improvement", "Down Payment Saving", "Debt-to-Income", "Pre-approval Process", "Closing Costs"]
            )
    
    with col2:
        tone = st.selectbox(
            "Content Tone:",
            ["Professional", "Friendly & Casual", "Educational", "Urgent/Action-Oriented", "Celebratory"]
        )
        
        include_cta = st.checkbox("Include Call-to-Action", value=True)
        include_hashtags = st.checkbox("Include Hashtags", value=True)
    
    # Personal/Company info
    with st.expander("📝 Personalization (Optional)"):
        loan_officer_name = st.text_input("Loan Officer Name", placeholder="John Smith")
        company_name = st.text_input("Company Name", placeholder="MyMCMB Mortgage")
        phone_number = st.text_input("Phone Number", placeholder="(555) 123-4567")
        website = st.text_input("Website", placeholder="www.mymcmb.com")
    
    def generate_social_content(content_type, platforms, **kwargs):
        """Generate social media content using AI"""
        try:
            # Build platform-specific requirements
            platform_specs = {
                "📘 Facebook": "Facebook posts should be engaging, can be longer (up to 500 characters recommended), and work well with questions and community engagement.",
                "📸 Instagram": "Instagram posts should be visual-focused, shorter text (125-150 characters optimal), include relevant hashtags, and work well with emojis.",
                "🐦 Twitter/X": "Twitter/X posts must be under 280 characters, concise, punchy, and include relevant hashtags.",
                "💼 LinkedIn": "LinkedIn posts should be professional, educational, can be longer, and focus on industry insights and expertise.",
                "🎵 TikTok": "TikTok content should be trendy, engaging, use popular format ideas, and include relevant hashtags for discovery."
            }
            
            platform_requirements = "\n".join([platform_specs[p] for p in platforms if p in platform_specs])
            
            # Build personalization
            personal_info = ""
            if kwargs.get('loan_officer_name'):
                personal_info += f"Loan Officer: {kwargs['loan_officer_name']}\n"
            if kwargs.get('company_name'):
                personal_info += f"Company: {kwargs['company_name']}\n"
            if kwargs.get('phone_number'):
                personal_info += f"Phone: {kwargs['phone_number']}\n"
            if kwargs.get('website'):
                personal_info += f"Website: {kwargs['website']}\n"
            
            prompt = f"""
            You are a social media content creator specializing in mortgage and real estate content for loan officers.
            
            CONTENT REQUEST:
            - Content Type: {content_type}
            - Platforms: {', '.join(platforms)}
            - Tone: {kwargs.get('tone', 'Professional')}
            - Include Call-to-Action: {kwargs.get('include_cta', True)}
            - Include Hashtags: {kwargs.get('include_hashtags', True)}
            
            PLATFORM REQUIREMENTS:
            {platform_requirements}
            
            PERSONALIZATION INFO:
            {personal_info if personal_info else "Generic content (no personalization provided)"}
            
            ADDITIONAL CONTEXT:
            """
            
            # Add context based on content type
            if "Market Update" in content_type or "Rate Alert" in content_type:
                prompt += f"Current 30-year rate: {kwargs.get('current_rate', 7.0)}%, Rate movement: {kwargs.get('rate_change', 'Stable')}\n"
            
            if "Educational" in content_type or "Tips" in content_type:
                prompt += f"Focus area: {kwargs.get('focus_area', 'General mortgage advice')}\n"
            
            prompt += f"""
            
            Please create engaging social media content that:
            1. Is appropriate for each selected platform
            2. Follows the specified tone and style
            3. Includes relevant mortgage/real estate information
            4. Incorporates personalization when provided
            5. Uses platform-appropriate formatting and length
            6. Includes effective hashtags if requested
            7. Has a clear call-to-action if requested
            
            Return a JSON object with the following structure:
            {{
                "posts": [
                    {{
                        "platform": "platform name",
                        "content": "the actual post content",
                        "hashtags": ["list", "of", "hashtags"],
                        "character_count": number,
                        "tips": "any additional tips for this platform"
                    }}
                ]
            }}
            """
            
            response = model.generate_content(
                prompt,
                generation_config=genai.types.GenerationConfig(
                    response_mime_type="application/json"
                )
            )
            
            return clean_json_response(response.text)
            
        except Exception as e:
            return {
                "posts": [],
                "error": f"Failed to generate content: {str(e)}"
            }
    
    # Generate content button
    if st.button("🚀 Generate Social Media Content", type="primary", use_container_width=True):
        with st.spinner("🤖 Creating engaging social media content..."):
            content_data = generate_social_content(
                content_type=content_type,
                platforms=platforms,
                tone=tone,
                include_cta=include_cta,
                include_hashtags=include_hashtags,
                current_rate=locals().get('current_rate'),
                rate_change=locals().get('rate_change'),
                focus_area=locals().get('focus_area'),
                loan_officer_name=loan_officer_name,
                company_name=company_name,
                phone_number=phone_number,
                website=website
            )
            
            if content_data.get('error'):
                st.error(content_data['error'])
            elif content_data.get('posts'):
                st.session_state.generated_content = content_data
                st.success("🎉 Content generated successfully!")
            else:
                st.error("No content was generated. Please try again.")
    
    # Display generated content
    if 'generated_content' in st.session_state and st.session_state.generated_content.get('posts'):
        st.markdown("---")
        st.header("📱 Generated Social Media Content")
        
        posts = st.session_state.generated_content['posts']
        
        # Create downloadable content
        all_content = []
        
        for i, post in enumerate(posts):
            platform = post.get('platform', f'Platform {i+1}')
            content = post.get('content', '')
            hashtags = post.get('hashtags', [])
            char_count = post.get('character_count', len(content))
            tips = post.get('tips', '')
            
            # Display each post
            with st.expander(f"{platform} Post ({char_count} characters)", expanded=True):
                # Content box
                st.text_area(
                    "Post Content:",
                    value=content,
                    height=150,
                    key=f"content_{i}",
                    help="Copy this content to your social media platform"
                )
                
                # Hashtags
                if hashtags:
                    hashtag_text = " ".join([f"#{tag}" if not tag.startswith('#') else tag for tag in hashtags])
                    st.text_area(
                        "Hashtags:",
                        value=hashtag_text,
                        height=60,
                        key=f"hashtags_{i}"
                    )
                
                # Platform tips
                if tips:
                    st.info(f"💡 **Platform Tip**: {tips}")
                
                # Character count indicator
                if char_count:
                    if platform == "🐦 Twitter/X" and char_count > 280:
                        st.warning(f"⚠️ Content is {char_count} characters (Twitter limit: 280)")
                    elif platform == "📸 Instagram" and char_count > 150:
                        st.info(f"ℹ️ Content is {char_count} characters (Instagram optimal: 125-150)")
                    else:
                        st.success(f"✅ {char_count} characters - Perfect length!")
            
            # Add to downloadable content
            all_content.append(f"""
PLATFORM: {platform}
CHARACTER COUNT: {char_count}

CONTENT:
{content}

HASHTAGS:
{hashtag_text if hashtags else 'No hashtags'}

TIPS:
{tips if tips else 'No specific tips'}

{'-'*80}
""")
        
        # Download options
        st.markdown("### 📥 Export Options")
        col1, col2 = st.columns(2)
        
        with col1:
            # Text export
            export_text = f"""
SOCIAL MEDIA CONTENT EXPORT
Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Content Type: {content_type}
Platforms: {', '.join(platforms)}
Tone: {tone}

{''.join(all_content)}
            """.strip()
            
            st.download_button(
                label="📄 Download as Text File",
                data=export_text,
                file_name=f"social_content_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                mime="text/plain"
            )
        
        with col2:
            # CSV export for content calendar
            csv_data = []
            for post in posts:
                csv_data.append({
                    'Platform': post.get('platform', ''),
                    'Content': post.get('content', ''),
                    'Hashtags': ' '.join(post.get('hashtags', [])),
                    'Character_Count': post.get('character_count', 0),
                    'Content_Type': content_type,
                    'Generated_Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
            
            csv_df = pd.DataFrame(csv_data)
            csv_buffer = io.StringIO()
            csv_df.to_csv(csv_buffer, index=False)
            
            st.download_button(
                label="📊 Download as CSV",
                data=csv_buffer.getvalue(),
                file_name=f"social_calendar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
        
        # Quick regenerate button
        if st.button("🔄 Generate New Content", type="secondary"):
            if 'generated_content' in st.session_state:
                del st.session_state.generated_content
            st.rerun()
    
    # Content inspiration section
    st.markdown("---")
    with st.expander("💡 Content Ideas & Best Practices"):
        st.markdown("""
        **🎯 High-Engagement Content Ideas:**
        
        **Market Updates:**
        - Weekly rate summaries with trend analysis
        - Local market insights and statistics
        - Economic factors affecting rates
        
        **Educational Content:**
        - "Mortgage Myth Monday" series
        - Step-by-step homebuying guides
        - Credit score improvement tips
        - Down payment strategies
        
        **Client Success Stories:**
        - Before/after scenarios (anonymized)
        - First-time buyer celebrations
        - Refinance success stories
        
        **Interactive Content:**
        - "Rate or Wait?" polls
        - Q&A sessions
        - Live market updates
        - Calculator tools and demos
        
        **📅 Posting Best Practices:**
        - **Facebook**: 1-2 posts per day, focus on community engagement
        - **Instagram**: 1 post + 2-3 stories daily, use high-quality visuals
        - **LinkedIn**: 3-5 posts per week, professional insights
        - **Twitter/X**: 2-5 tweets daily, join conversations with hashtags
        - **TikTok**: 3-5 videos per week, follow trends and use trending sounds
        
        **⏰ Optimal Posting Times:**
        - **Facebook**: 9 AM - 10 AM, 3 PM - 4 PM
        - **Instagram**: 6 AM - 9 AM, 7 PM - 9 PM  
        - **LinkedIn**: 8 AM - 10 AM, 12 PM - 2 PM
        - **Twitter/X**: 8 AM - 10 AM, 7 PM - 9 PM
        - **TikTok**: 6 AM - 10 AM, 7 PM - 9 PM
        """)
